import streamlit as st
import pandas as pd
import os
import csv
import io
import ollama
from datetime import datetime
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
import shutil
import zipfile
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, range_boundaries
from openpyxl.utils.cell import range_boundaries
from io import BytesIO

#streamlit run main.py
#ollama run mistral

st.set_page_config(
    page_title="Сервис проверки вопросов КРИ",
    page_icon="UrFULogo.png",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        "Get help": "https://t.me/ur_rinatmuginov",
        "Report a bug": "https://t.me/ur_rinatmuginov",
        "About": "ВКР. Мугинов Ринат. Приложение для проверки тестов для КРИ"
    }
)

st.sidebar.markdown("""
<h3 style='font-size: 24px;'>Меню</h3>
""", unsafe_allow_html=True)

section = st.sidebar.radio(
    label="",
    options=[
        "1. Вопросы для самопроверки",
        "2. Проверка ответов студентов",
        "3. Формирование итогового журнала",
        #"4. Формирование ведомостей и журналов"
    ],
    index=0,
    format_func=lambda x: x.upper(),
    disabled=False,
    horizontal=False,
    captions=[
        "Редактирование базы вопросов",
        "Автоматическая проверка через ИИ",
        "Генерация журнала на основании проверенных тестов",
        "Генератор пустых ведомостей и журналов"
    ],
    key="main_menu_radio"
)

st.sidebar.markdown("")
st.sidebar.markdown("---")
st.sidebar.markdown("""
<div style="text-align: center; font-size: 1em; color: #666;">
    Версия 1.0 | © 2025<br>
    <a href="https://t.me/ur_rinatmuginov" style="color: #666;">Сообщить об ошибке</a>
</div>
""", unsafe_allow_html=True)

if section == "1. Вопросы для самопроверки":

    def is_valid_lecture_id(value):
        return bool(re.fullmatch(r"Lec\d{2}", str(value)))

    def is_valid_question_id(value):
        return bool(re.fullmatch(r"Q\d{3}", str(value)))

    st.title("Редактор базы вопросов для самопроверки")

    DB_PATH = "data/base_questions.xlsx"

    if 'df_ethalons' not in st.session_state:
        try:
            if os.path.exists(DB_PATH):
                st.session_state.df_ethalons = pd.read_excel(DB_PATH)
                st.success("База эталонов загружена!")
            else:
                st.session_state.df_ethalons = pd.DataFrame(columns=[
                    "Discipline", "Lecture_ID", "Question_ID", "Question", "Answer"
                ])
                st.warning("Файл не найден. Создана новая база.")
        except Exception as e:
            st.error(f"Ошибка загрузки: {str(e)}")
            st.session_state.df_ethalons = pd.DataFrame()

    df = st.session_state.df_ethalons

    st.subheader("Фильтрация вопросов")
    available_disciplines = df["Discipline"].dropna().unique()
    selected_discipline = st.selectbox("Выберите дисциплину", options=["Все"] + list(available_disciplines))

    available_sessions = df["Lecture_ID"].dropna().unique()  # Проверь, может тоже "Session"?
    selected_sessions = st.multiselect(
        "Выберите номер лекции",
        options=available_sessions,
        placeholder="Например: Lec01, Lec02"
    )

    filtered_df = df.copy()
    if selected_discipline != "Все":
        filtered_df = filtered_df[filtered_df["Discipline"] == selected_discipline]
    if selected_sessions:
        filtered_df = filtered_df[filtered_df["Lecture_ID"].isin(selected_sessions)]

    st.subheader("Таблица вопросов")
    st.caption(f"Отображается {len(filtered_df)} вопрос(ов)")

    edited_df = st.data_editor(filtered_df, num_rows="dynamic", use_container_width=True)

    def download_excel(df):
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        return output.getvalue()

    excel_bytes = download_excel(edited_df)
    discipline_part = selected_discipline if selected_discipline != "Все" else "Все_дисциплины"
    session_part = "_".join(map(str, selected_sessions)) if selected_sessions else "все_занятия"
    filename = f"вопросы_{discipline_part}_{session_part}.xlsx".replace(" ", "_")

    st.download_button(
        "Скачать отфильтрованную таблицу",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if st.button("Сохранить отредактированные данные в файл"):
        try:
            # Валидация
            invalid_lecture_ids = edited_df[~edited_df["Lecture_ID"].apply(is_valid_lecture_id)]
            invalid_question_ids = edited_df[~edited_df["Question_ID"].apply(is_valid_question_id)]

            if not invalid_lecture_ids.empty or not invalid_question_ids.empty:
                st.error("❌ Обнаружены некорректные значения:")
                if not invalid_lecture_ids.empty:
                    st.write("Неверный Lecture_ID:", invalid_lecture_ids[["Lecture_ID"]])
                if not invalid_question_ids.empty:
                    st.write("Неверный Question_ID:", invalid_question_ids[["Question_ID"]])
                st.stop()

            # Работаем с полной таблицей
            full_df = st.session_state.df_ethalons

            # Ключ: Discipline + Lecture_ID + Question_ID
            key_cols = ["Discipline", "Lecture_ID", "Question_ID"]

            # Определим строки, которые были отфильтрованы, но удалены пользователем
            filtered_keys = filtered_df[key_cols].apply(tuple, axis=1)
            edited_keys = edited_df[key_cols].apply(tuple, axis=1)
            removed_keys = set(filtered_keys) - set(edited_keys)

            # Удалим удалённые строки из полной таблицы
            full_df = full_df[~full_df[key_cols].apply(tuple, axis=1).isin(removed_keys)]

            # Обновим и добавим строки
            full_df.set_index(key_cols, inplace=True)
            edited_df.set_index(key_cols, inplace=True)
            full_df.update(edited_df)

            # Добавим новые строки (если есть)
            new_rows = edited_df[~edited_df.index.isin(full_df.index)]
            full_df = pd.concat([full_df, new_rows])

            # Сохраняем
            full_df.reset_index(inplace=True)
            st.session_state.df_ethalons = full_df
            full_df.to_excel(DB_PATH, index=False)
            st.success("✅ Изменения успешно сохранены.")

        except Exception as e:
            st.error(f"Ошибка при сохранении: {e}")

elif section == "2. Проверка ответов студентов":

    ETHALON_PATH = "data/base_questions.xlsx"
    ID_IDENTIFIER = "Укажите Ваш ID"
    LLM_MODEL = "mistral"

    st.title("Проверка ответов студентов")

    uploaded_file = st.file_uploader(
        "Перетащите CSV-файл с ответами студентов из сервиса wj.qq",
        type=["csv"]
    )

    if uploaded_file:
        try:
            text_io = io.TextIOWrapper(uploaded_file, encoding="utf-8", newline='')
            csv_df = pd.read_csv(text_io, engine="python")

            ethalons_df = pd.read_excel(ETHALON_PATH)
            disciplines = ethalons_df["Discipline"].dropna().unique()
            selected_discipline = st.selectbox("Выберите дисциплину", disciplines)

            if selected_discipline:
                lecture_ids = ethalons_df[ethalons_df["Discipline"] == selected_discipline]["Lecture_ID"].unique()
                selected_lecture = st.selectbox("Выберите лекцию", lecture_ids)

                if selected_lecture and st.button("Проверить тесты"):
                    id_column = next((col for col in csv_df.columns if ID_IDENTIFIER in str(col)), None)
                    if not id_column:
                        st.error("Не найдена колонка с ID студентов!")
                        st.stop()

                    question_start_index = None
                    for i, col in enumerate(csv_df.columns):
                        if "Как называется" in col or col.strip().startswith("1."):
                            question_start_index = i
                            break
                    if question_start_index is None:
                        st.error("Не удалось найти начало вопросов в CSV!")
                        st.stop()

                    ethalons = ethalons_df[
                        (ethalons_df["Discipline"] == selected_discipline) &
                        (ethalons_df["Lecture_ID"] == selected_lecture)
                        ].sort_values("Question_ID")

                    ethalon_questions = ethalons["Question"].astype(str).tolist()
                    ethalon_answers = ethalons["Answer"].astype(str).tolist()
                    num_questions = len(ethalon_questions)

                    def grade_with_llm(question, reference_answer, student_answer):
                        if not student_answer.strip():
                            return 0, "Пустой ответ"
                        prompt = f"""
    Ты — преподаватель. Проверь, насколько ответ студента совпадает с эталонным по смыслу. Оценивай нестрого. Полное соответствие необязательно. 

    Вопрос: {question}
    Эталонный ответ: {reference_answer}
    Ответ студента: {student_answer}

    Оцени по шкале:
    - 0 — не по теме
    - 1 — частично верно
    - 2 — полностью верно, но могут быть недочеты

    Ответь только числом: 0, 1 или 2.
    """
                        try:
                            response = ollama.chat(model=LLM_MODEL, messages=[{"role": "user", "content": prompt}])
                            reply = response["message"]["content"].strip()
                            score = int([s for s in reply.split() if s.isdigit()][0])
                            return min(max(score, 0), 2), reply
                        except Exception as e:
                            return 0, f"Ошибка: {e}"

                    results = []
                    logs = []
                    detailed_results = []

                    for idx, row in csv_df.iterrows():
                        student_id = str(row[id_column]).strip()
                        student_answers_raw = row.values[question_start_index:question_start_index + num_questions]
                        student_answers = [str(ans).strip() if pd.notna(ans) else "" for ans in student_answers_raw]

                        total_score = 0
                        student_detailed_results = []

                        with ThreadPoolExecutor(max_workers=10) as executor:
                            futures = {
                                executor.submit(grade_with_llm, question, correct_ans, student_ans): (
                                i + 1, question, correct_ans, student_ans)
                                for i, (question, correct_ans, student_ans) in
                                enumerate(zip(ethalon_questions, ethalon_answers, student_answers))
                            }

                            for future in as_completed(futures):
                                i, question, correct_ans, student_ans = futures[future]
                                score, _ = future.result()
                                total_score += score

                                student_detailed_results.append({
                                    "ID студента": student_id,
                                    "Вопрос №": i,
                                    "Вопрос": question,
                                    "Эталонный ответ": correct_ans,
                                    "Ответ студента": student_ans,
                                    "Балл": score
                                })

                        results.append({
                            "ID студента": student_id,
                            f"Оценка (из 10)": total_score
                        })
                        detailed_results.extend(student_detailed_results)

                    log_df = pd.DataFrame(detailed_results)
                    results_df = pd.DataFrame(results)

                    temp_dir = "temp"
                    os.makedirs(temp_dir, exist_ok=True)

                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    safe_discipline = re.sub(r'\W+', '_', selected_discipline)
                    safe_lecture = re.sub(r'\W+', '_', selected_lecture)

                    results_filename = f"{temp_dir}/results_{safe_discipline}_{safe_lecture}_{timestamp}.csv"
                    log_filename = f"{temp_dir}/log_{safe_discipline}_{safe_lecture}_{timestamp}.csv"
                    results_df.to_csv(results_filename, index=False)
                    log_df.to_csv(log_filename, index=False)

                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        # Добавляем файлы в архив, читая их обратно из диска
                        with open(results_filename, 'rb') as f:
                            zipf.writestr(f"results_{safe_discipline}_{safe_lecture}_{timestamp}.csv", f.read())
                        with open(log_filename, 'rb') as f:
                            zipf.writestr(f"log_{safe_discipline}_{safe_lecture}_{timestamp}.csv", f.read())

                    st.download_button(
                        label="Скачать все результаты (ZIP)",
                        data=zip_buffer.getvalue(),
                        file_name=f"results_{safe_discipline}_{safe_lecture}_{timestamp}.zip",
                        mime="application/zip"
                    )

                    st.subheader("Итоговая таблица:")
                    st.dataframe(results_df, use_container_width=True)

                    if "llm_results" not in st.session_state:
                        st.session_state["llm_results"] = results_df

                    if "llm_logs" not in st.session_state:
                        st.session_state["llm_logs"] = log_df

        except Exception as e:
            st.error(f"Ошибка при загрузке файла: {e}")

elif section == "3. Формирование итогового журнала":

    st.title("Загрузка результатов в журнал")

    journal_dir = "data/journals"
    journal_files = [f for f in os.listdir(journal_dir) if f.endswith(".xlsx")]
    selected_journal_file = st.selectbox("Выберите файл журнала (.xlsx)", journal_files)

    if selected_journal_file:
        journal_path = os.path.join(journal_dir, selected_journal_file)
        st.success(f"Выбран файл журнала: {selected_journal_file}")

        result_files = [f for f in os.listdir("temp") if f.startswith("results_") and f.endswith(".csv")]
        selected_result_file = st.selectbox("Выберите файл с результатами (.csv)", result_files)

        if selected_result_file:
            st.success(f"Выбран файл результатов: {selected_result_file}")
            results_path = os.path.join("temp", selected_result_file)

            df_results = pd.read_csv(results_path, header=None, names=["ID", "Баллы"])

            match = re.match(r'results_(.*?)_Lec(\d+)_', selected_result_file)
            if match:
                discipline_name = match.group(1)
                lecture_number = int(match.group(2))
                st.write(f"Дисциплина: **{discipline_name}**")
                st.write(f"Номер лекции: **{lecture_number}**")

                workbook = load_workbook(journal_path)
                sheet = workbook.active

                columns = ['F', 'H', 'J', 'L', 'N', 'P', 'R', 'T', 'V', 'X']
                if lecture_number > len(columns):
                    st.error("Слишком большой номер лекции. В шаблоне нет такой колонки.")
                else:
                    target_col = columns[lecture_number - 1]
                    id_column = 'B'
                    max_row = sheet.max_row

                    id_to_score = {str(k): v for k, v in zip(df_results["ID"], df_results["Баллы"])}
                    updated_count = 0
                    found_ids = set()

                    for row in range(8, max_row + 1):
                        student_id = str(sheet[f"{id_column}{row}"].value)
                        if student_id in id_to_score:
                            col_idx = column_index_from_string(target_col)

                            for merged_range in sheet.merged_cells.ranges:
                                min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                                if min_row <= row <= max_row and min_col <= col_idx <= max_col:
                                    cell = sheet.cell(row=min_row, column=min_col)
                                    break
                            else:
                                cell = sheet.cell(row=row, column=col_idx)

                            cell.value = id_to_score[student_id]
                            updated_count += 1
                            found_ids.add(student_id)

                    missing_ids = set(map(str, df_results["ID"])) - found_ids
                    if missing_ids:
                        st.warning(f"⚠️ В журнале не найдены ID {len(missing_ids)} студентов.")
                        st.write("Вот их список:")
                        st.code("\n".join(missing_ids))
                    else:
                        st.info("✅ Все студенты из результатов найдены в журнале.")

                    st.success(f"Обновлено {updated_count} записей.")

                    data = []
                    for row in sheet.iter_rows(values_only=True):
                        data.append(row)
                    df_preview = pd.DataFrame(data)
                    st.dataframe(df_preview)

                    def to_excel(wb):
                        output = BytesIO()
                        wb.save(output)
                        return output.getvalue()

                    excel_data = to_excel(workbook)
                    st.download_button(
                        label="Скачать обновленный журнал",
                        data=excel_data,
                        file_name=f"обновленный_журнал_{discipline_name}_lec{lecture_number}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.error("Не удалось определить название дисциплины и номер лекции из имени файла.")

#elif section == "4. Формирование ведомостей и журналов":

    # === Вспомогательные функции ===
    def clean_filename(name):
        return re.sub(r'[<>:"/\\|?*_]', '', str(name)).strip()

    def determine_semester_and_year():
        now = datetime.now()
        year = now.year
        semester = "весенний" if now.month <= 6 else "осенний"
        return semester, year

    #def create_documents(database_path, output_folder):
        logs = []
        try:
            data = pd.read_excel(database_path)
        except Exception as e:
            return [f"Ошибка при чтении файла базы данных: {e}"]

        os.makedirs(output_folder, exist_ok=True)

        template_prefixes = {
            "Арх": "EXAMPLE_A",
            "Гео": "EXAMPLE_G",
            "Строй": "EXAMPLE_S",
            "ТиТ": "EXAMPLE_T",
            "ВиВ": "EXAMPLE_V"
        }

        direction_folders = {
            "Арх": "Архитектура",
            "Гео": "Геодезия",
            "Строй": "Строительство",
            "ТиТ": "ТиТ",
            "ВиВ": "ВиВ"
        }

        current_semester, current_year = determine_semester_and_year()

        for index, row in data.iterrows():
            try:
                direction = str(row.get('Направление', '')).strip()
                discipline = str(row.get('Название дисциплины', '')).strip()
                teacher = str(row.get('Преподаватель', '')).strip()
                course_year = str(row.get('Курс', '')).strip()
                hours = str(row.get('Часы', '')).strip()

                if not all([direction, discipline, teacher, course_year, hours]):
                    logs.append(f"Пропущена строка {index + 1}: не все данные заполнены.")
                    continue

                template_prefix = template_prefixes.get(direction)
                if not template_prefix:
                    logs.append(f"Для направления '{direction}' не определен шаблон ведомости. Пропускаю строку.")
                    continue

                statement_template_file = os.path.join(os.path.dirname(database_path),
                                                       f"V_{template_prefix}{course_year}.xlsx")
                if not os.path.exists(statement_template_file):
                    logs.append(f"Файл шаблона ведомости не найден: {statement_template_file}")
                    continue

                workbook = load_workbook(statement_template_file)
                sheet = workbook.active
                sheet['A2'] = discipline

                last_row = max([cell.row for cell in sheet['C'] if cell.value is not None], default=1)
                teacher_row = last_row + 2
                sheet[f'C{teacher_row}'] = teacher
                sheet[f'C{teacher_row}'].font = Font(name='Times New Roman', bold=True)

                direction_folder = direction_folders.get(direction)
                if not direction_folder:
                    logs.append(f"Для направления '{direction}' не определена папка. Пропускаю строку.")
                    continue

                statements_output_folder = os.path.join(output_folder, "ведомости", direction_folder)
                os.makedirs(statements_output_folder, exist_ok=True)
                discipline_clean = clean_filename(discipline)
                statement_output_file = os.path.join(statements_output_folder,
                                                     f"Ведомость {discipline_clean} {course_year}.xlsx")
                workbook.save(statement_output_file)

                # Журнал
                journal_template_file = os.path.join(os.path.dirname(database_path), f"J_EXAMPLE_A2022.xlsx")
                if not os.path.exists(journal_template_file):
                    logs.append(f"Файл шаблона журнала не найден: {journal_template_file}")
                    continue

                workbook = load_workbook(journal_template_file)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet['A2'] = discipline
                    sheet['A4'] = f"{current_semester.capitalize()} семестр {current_year}\nКоличество часов: {hours}"

                    last_row = max([cell.row for cell in sheet['D'] if cell.value is not None], default=1)
                    teacher_row = last_row + 2
                    sheet[f'D{teacher_row}'] = teacher
                    sheet[f'D{teacher_row}'].font = Font(name='Times New Roman', bold=True)

                journals_output_folder = os.path.join(output_folder, "Журналы", direction_folder)
                os.makedirs(journals_output_folder, exist_ok=True)
                journal_output_file = os.path.join(journals_output_folder,
                                                   f"Журнал {discipline_clean} {course_year}.xlsx")
                workbook.save(journal_output_file)

            except Exception as e:
                logs.append(f"Ошибка при обработке строки {index + 1}: {e}")

        logs.append("Все документы успешно созданы.")
        return logs